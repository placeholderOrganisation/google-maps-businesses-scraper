from modules.const.settings import SETTINGS
import openpyxl

def write_to_workbook(data, row, col, workbook_name=SETTINGS["WORKBOOK_NAME"], worksheet_name=SETTINGS["WORKSHEET_NAME"]):
    # Open the Excel file
    workbook = openpyxl.load_workbook(workbook_name)
    try:
        # Select the first sheet
        sheet = workbook.get_sheet_by_name(worksheet_name)
        # Write data to row 4, column 5
        sheet.cell(row=row, column=col).value = data
    finally:
        workbook.save(workbook_name)

def count_entries_in_workbook(workbook_name=SETTINGS["WORKBOOK_NAME"]):
    # Open the Excel file
    workbook = openpyxl.load_workbook(workbook_name)
    try:
        # Select the first sheet
        sheet = workbook.worksheets[0]
        # Count the number of rows in the sheet
        num_rows = sheet.max_row
    finally:
        workbook.close()
        # return the result
        return num_rows

def extract_column_from_row(row, column, workbook_name=SETTINGS["WORKBOOK_NAME"]):
    # Open the Excel file
    workbook = openpyxl.load_workbook(workbook_name)
    try:
        # Select the first sheet
        sheet = workbook.worksheets[0]
        # Extract value from column 5
        column_value = sheet.cell(row=row, column=column).value
    finally:
        return column_value